from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportExportService:
    """Экспорт табличных отчётов в PDF и Excel."""

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _to_table_data(self, data: list[dict], title: str) -> tuple[list[str], list[list[Any]]]:
        """Преобразует список словарей в заголовки и строки таблицы."""
        if not data:
            return [], []
        headers = list(data[0].keys())
        rows = [[row.get(h, "") for h in headers] for row in data]
        return headers, rows

    def export_pdf(self, data: list[dict], title: str, filename: str | None = None) -> str:
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("Библиотека reportlab не установлена. Установите: pip install reportlab")

        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = self.output_dir / filename

        headers, rows = self._to_table_data(data, title)
        if not headers:
            raise ValueError("Нет данных для экспорта")

        doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,  # center
            spaceAfter=12
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Таблица
        table_data = [headers] + rows
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E2F3')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        doc.build(elements)
        return str(filepath)

    def export_excel(self, data: list[dict], title: str, filename: str | None = None) -> str:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Библиотека openpyxl не установлена. Установите: pip install openpyxl")

        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename

        headers, rows = self._to_table_data(data, title)
        if not headers:
            raise ValueError("Нет данных для экспорта")

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Заголовок листа
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

        # Заголовки таблицы
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Данные
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(4, len(rows) + 5):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_length = max(max_length, len(str(val)))
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = min(max_length + 4, 50)

        wb.save(str(filepath))
        return str(filepath)